from tkcalendar import Calendar
from tkinter import filedialog
from tkinter import font
from openpyxl.styles import colors, PatternFill, Font, Color
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from threading import Thread
from decimal import Decimal
import tkinter as tk
import subprocess
import requests
import shutil
import json
import time
import os

BTC_SYMBOL = 'BTC'
BTC_DECIMAL = 1e8

HISTORICAL_PRICE_DATA_OUTPUT_PATH = 'historical_price_data'
SENT_RECEIVED_HISTORY_OUTPUT_PATH = 'sent_received_history'
VERIFIED_TRANSACTIONS_OUTPUT_PATH = 'verified_transactions'

btcusd_history = None

def verify_address():
    address_status.set('Verifying Address...')
    address = address_entry.get().strip()
    is_valid = requests.get(f'https://mempool.space/api/address/{address}')
    
    copy_address_status.set('')
    
    if is_valid.status_code == 400:
        address_status.set('Invalid Address')
    
    content = json.loads(is_valid.content)
    
    address_status.set(f'https://mempool.space/address/{address}')

def verify_tx():
    confirm_blockchain.set('Verifying Transaction...')
    tx_hash = tx_hash_entry.get().strip()
    is_valid = requests.get(f'https://mempool.space/api/tx/{tx_hash}')
    
    copy_tx_status.set('')
    
    if is_valid.status_code == 400:
        confirm_blockchain.set('Invalid Transaction')
        return
    
    confirm_blockchain.set(f'https://mempool.space/tx/{tx_hash}')

def tx_exists(tx_hash):
    is_valid = requests.get(f'https://mempool.space/api/tx/{tx_hash.strip()}')
    
    if is_valid.status_code == 400:
        return False
    
    return True   
    
def get_current_value():
    price = requests.get('https://api.cryptowat.ch/markets/bitfinex/btcusd/price')
    content = json.loads(price.content)
    return Decimal(content['result']['price'])
    
def get_entry_price(tx_time):
    days = tx_time // 86400
    tx_time = days * 86400
    for row in btcusd_history:
        if row[0] == tx_time:
            return Decimal(row[4])
    
    return 1

def get_txs(address):
    startdate = int(datetime.strptime(start_date.get(), "%m/%d/%y").timestamp())
    enddate = int(datetime.strptime(end_date.get(), "%m/%d/%y").timestamp())
    
    filtered_txs = []

    last_seen_txid = ''
    last_seen_txid = get_txs_chain(address, startdate, enddate, filtered_txs, last_seen_txid)
    while last_seen_txid:
        last_seen_txid = get_txs_chain(address, startdate, enddate, filtered_txs, last_seen_txid)
        
    return filtered_txs

def get_total_value(address, funds):
    total = 0
    for fund in funds:
        if 'prevout' in fund:
            fund = fund['prevout']
        if 'scriptpubkey_address' in fund:
            fund_address = fund['scriptpubkey_address']
            if address == fund_address:
                total += int(fund['value'])
    
    return total
    
def round_2_decimal_places(amount):
    amount *= 100
    amount = round(amount)
    return amount / 100
    
def get_txs_chain(address, startdate, enddate, filtered_txs, last_seen_txid):
    txs = requests.get(f'https://mempool.space/api/address/{address}/txs/chain/{last_seen_txid}')
    
    content = json.loads(txs.content)
    
    if txs.status_code == 400:
        return ''
        
    current_value = get_current_value()
    
    last_tx = ''
    for tx in content:
        tx_time = int(tx['status']['block_time']) - 28800
        
        if tx_time < startdate:
            break

        last_tx = tx['txid']

        if tx_time > enddate:
            continue

        sent = get_total_value(address, tx['vin'])/BTC_DECIMAL
        received = get_total_value(address, tx['vout'])/BTC_DECIMAL
        
        if sent != 0 and received != 0:
            if sent > received:
                sent -= received
                received = 0
            else:
                received -= sent
                sent = 0
            
        tx_data = [datetime.fromtimestamp(tx_time).strftime('%Y-%m-%d %H:%M:%S'),
                   tx['status']['block_height'], tx['txid'], sent, received, BTC_SYMBOL]

        if include_ugl.get():
            value = sent if sent != 0 else received
            tx_data.append(f'${round_2_decimal_places(current_value * Decimal(value))}')
            entry_price = get_entry_price(tx_time + 57600)
            tx_data.append(f'${round_2_decimal_places(entry_price * Decimal(value))}')
            
        filtered_txs.append(tx_data)

    return last_tx

def generate_sent_received_history():
    send_received_status.set('Retrieving Transactions...')
    
    address = sent_received_address.get().strip()    
    rows = get_txs(address)
    
    filename = f'{SENT_RECEIVED_HISTORY_OUTPUT_PATH}/{BTC_SYMBOL}_{address}.xlsx'
    
    work_book = Workbook()
    work_sheet = work_book.active
    columns = ['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset']
    
    if include_ugl.get():
        columns.append('Current USD Value (@currentprice * amount)')
        columns.append('Entry Price Value(USD)')

    work_sheet.append(columns)
    
    for row in rows:
        work_sheet.append(row)
        
    work_book.save(filename)

    send_received_status.set(f'Transactions retrieved successfully as of {datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")}')

def get_history(before, after):
    before += 28800
    market = current_market.get()
    history = requests.get(f'https://api.cryptowat.ch/markets/{market}/btcusd/ohlc?before={before}&after={after}&periods=86400')
    content = json.loads(history.content)
    return content['result']['86400']


def generate_historical_price_data():
    history_status.set('Retrieving BTC historical price data...')
    
    before = int(datetime.strptime(history_end_date.get(), "%m/%d/%y").timestamp())
    after = int(datetime.strptime(history_start_date.get(), "%m/%d/%y").timestamp())

    rows = get_history(before, after)
    
    filename = f'{HISTORICAL_PRICE_DATA_OUTPUT_PATH}/{BTC_SYMBOL}_price_history_{after}-{before}.xlsx'
    
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.append(['Date', 'Closing Price(USD)'])

    for row in rows:
        date = datetime.fromtimestamp(int(row[0])).strftime('%m/%d/%Y')
        work_sheet.append([date, row[4]])        
    
    work_book.save(filename)

    history_status.set(f'BTC historical price data retrieved successfully')    

def verify_multi_tx():
    file = filedialog.askopenfilename()
    shutil.copy(file, VERIFIED_TRANSACTIONS_OUTPUT_PATH)
    filename = os.path.basename(file)
    file = f'{VERIFIED_TRANSACTIONS_OUTPUT_PATH}/{filename}'
    work_book = load_workbook(file, data_only=True)
    
    upload_status.set(f'Working with {filename}')

    for worksheet_idx in range(len(work_book.worksheets)):
        tx_id_idx = None
        bc_url_idx = None
        work_book.active = worksheet_idx
        work_sheet = work_book.active
        upload_status.set(f'Working with sheet {work_sheet.title}')
       
        for column_idx, column_name in enumerate(work_sheet[1]):
            if column_name.value in ('Transaction Details', 'Transaction ID', 'Transaction Detail'):
                tx_id_idx = column_idx
            elif column_name.value == 'Blockchain URL':
                bc_url_idx = column_idx


        row_count = work_sheet.max_row
       
        for row_idx in range(2, row_count + 1):
            upload_status.set(f'Working with row #{row_idx}/{row_count}')
            if tx_id_idx is not None and bc_url_idx is not None:
                tx_id = work_sheet[row_idx][tx_id_idx].value
                if tx_exists(tx_id):
                    bc_link = f'https://mempool.space/tx/{tx_id}'
                    work_sheet[row_idx][bc_url_idx].value = '=HYPERLINK("{}", "{}")'.format(bc_link, 'Verified')
                    work_sheet[row_idx][bc_url_idx].fill = PatternFill(start_color='C3ECCB', end_color='C3ECCB', fill_type = 'solid')
                    work_sheet[row_idx][bc_url_idx].font = Font(color='006100')
                else:
                    work_sheet[row_idx][bc_url_idx].value = 'Null'
                    work_sheet[row_idx][bc_url_idx].fill = PatternFill(start_color='F2D3D7', end_color='F2D3D7', fill_type = 'solid')
                    work_sheet[row_idx][bc_url_idx].font = Font(color='9C0039')

    work_book.save(file)                        
    upload_status.set('Transactions Verified')

def set_start_date():
    start_date.set(calendar.get_date())

def set_end_date():
    end_date.set(calendar.get_date())
    
def copy_address():
    cmd = f'echo {address_status.get().strip()} | clip'
    subprocess.check_call(cmd, shell=True)
    copy_address_status.set('copied to clipboard')

def copy_tx():
    cmd = f'echo {confirm_blockchain.get().strip()} | clip'
    subprocess.check_call(cmd, shell=True)
    copy_tx_status.set('copied to clipboard')
    
def run_thread(process):
    Thread(target=process).start()

def init():
    exchange_history = requests.get(f'https://api.cryptowat.ch/markets/bitfinex/btcusd/ohlc?&periods=86400')
    content = json.loads(exchange_history.content)
    global btcusd_history
    btcusd_history = content['result']['86400']

init()
####
#UI#
####
window = tk.Tk()
window.title('')
window.geometry('1350x750')
window.iconbitmap('res/python-logo.ico')
window.configure(bg='#ffffff')

lato_font = font.Font(family='Lato')

tk.Label(window, text="Open Bitcoin for Excel", font=f'{lato_font} 24 bold', fg='#6030b1', bg='#ffffff').place(x=10, y=10)

copy_img = tk.PhotoImage(file="res/copy.png")

tk.Label(window, text="Address Checker", font=f'{lato_font} 18', fg='#6030b1', bg='#ffffff').place(x=20, y=60)

tk.Label(window, text="Address", font=f'{lato_font} 14', fg='#242424', bg='#ffffff').place(x=30, y=100)

address_entry = tk.Entry(width=90, borderwidth=0, font=f'{lato_font} 14', bg='#f8f8f8') 
address_entry.place(x=120, y=100)

verify_address_btn = tk.Button(text="   Verify   ", borderwidth=0, font=f'{lato_font} 14', bg='#6030b1', fg='#ffffff', command=lambda:run_thread(verify_address))
verify_address_btn.place(x=1120, y=95)

tk.Label(window, text="URL", font=f'{lato_font} 14', fg='#242424', bg='#ffffff').place(x=30, y=140)

address_status = tk.StringVar()
address_status_entry = tk.Entry(width=90, borderwidth=0, font=f'{lato_font} 14', bg='#f8f8f8', textvariable=address_status) 
address_status_entry.place(x=120, y=140)

copy_address_btn = tk.Button(border=0, bg='#ffffff', image=copy_img, command=lambda:run_thread(copy_address))
copy_address_btn.place(x=1120, y=145)

copy_address_status = tk.StringVar()
copy_address_txt = tk.Entry(width=32, borderwidth=0, font=f'{lato_font} 12 italic', bg='#ffffff', textvariable=copy_address_status)
copy_address_txt.place(x=1140, y=140) 


tk.Label(window, text="Transaction Verification", font=f'{lato_font} 18', fg='#6030b1', bg='#ffffff').place(x=20, y=190)

upload_img = tk.PhotoImage(file="res/upload_img.png")
upload_btn = tk.Button(border=0, image=upload_img, bg='#ffffff', command=lambda:run_thread(verify_multi_tx))
upload_btn.place(x=80, y=230)

tk.Label(window, text="Transaction Hash", font=f'{lato_font} 14', fg='#242424', bg='#ffffff').place(x=300, y=240)

tx_hash_entry = tk.Entry(width=60, borderwidth=0, font=f'{lato_font} 14', bg='#f8f8f8') 
tx_hash_entry.place(x=460, y=240)

verify_tx_btn = tk.Button(text="   Verify   ", borderwidth=0, font=f'{lato_font} 14', bg='#6030b1', fg='#ffffff', command=lambda:run_thread(verify_tx))
verify_tx_btn.place(x=1120, y=235)

upload_status = tk.StringVar()
upload_status_txt = tk.Entry(width=20, borderwidth=0, font=f'{lato_font} 12 italic', bg='#ffffff', textvariable=upload_status) 
upload_status_txt.place(x=70, y=275)

tk.Label(window, text="Blockchain Confirmation", font=f'{lato_font} 14', fg='#242424', bg='#ffffff').place(x=300, y=275)

confirm_blockchain = tk.StringVar()
confirm_blockchain_entry = tk.Entry(width=54, borderwidth=0, font=f'{lato_font} 14', bg='#f8f8f8', textvariable=confirm_blockchain) 
confirm_blockchain_entry.place(x=520, y=275)

copy_tx_btn = tk.Button(border=0, bg='#ffffff', image=copy_img, command=lambda:run_thread(copy_tx))
copy_tx_btn.place(x=1120, y=280)

copy_tx_status = tk.StringVar()
copy_tx_txt = tk.Entry(width=32, borderwidth=0, font=f'{lato_font} 12 italic', bg='#ffffff', textvariable=copy_tx_status)
copy_tx_txt.place(x=1140, y=280) 

tk.Label(window, text="Sent/Received History", font=f'{lato_font} 18', fg='#6030b1', bg='#ffffff').place(x=20, y=325)

calendar_img = tk.PhotoImage(file="res/calendar_icon.png")
start_date_calendar = Calendar(window, selectmode = 'day', year = 2020, month = 1, day = 1)
end_date_calendar = Calendar(window, selectmode = 'day', year = 2022, month = 1, day = 1)
history_start_date_calendar = Calendar(window, selectmode = 'day', year = 2020, month = 1, day = 1)
history_end_date_calendar = Calendar(window, selectmode = 'day', year = 2022, month = 1, day = 1)


def set_start_date():
    if start_date_calendar.winfo_ismapped():
        start_date.set(start_date_calendar.get_date())
        start_date_calendar.place_forget()
    else:
        start_date_calendar.place(x=130, y=375)
        start_date_calendar.tkraise()

def set_end_date():
    if end_date_calendar.winfo_ismapped():
        end_date.set(end_date_calendar.get_date())
        end_date_calendar.place_forget()
    else:
        end_date_calendar.place(x=600, y=375)
        end_date_calendar.tkraise()
        
def set_history_start_date():
    if history_start_date_calendar.winfo_ismapped():
        history_start_date.set(history_start_date_calendar.get_date())
        history_start_date_calendar.place_forget()
    else:
        history_start_date_calendar.place(x=130, y=555)
        history_start_date_calendar.tkraise()
        
def set_history_end_date():
    if history_end_date_calendar.winfo_ismapped():
        history_end_date.set(history_end_date_calendar.get_date())
        history_end_date_calendar.place_forget()
    else:
        history_end_date_calendar.place(x=600, y=555)
        history_end_date_calendar.tkraise()

tk.Label(window, text="Start Date", font=f'{lato_font} 14', fg='#242424', bg='#ffffff').place(x=30, y=375)

start_date = tk.StringVar()
start_date_entry = tk.Entry(width=30, borderwidth=0, font=f'{lato_font} 14', bg='#f8f8f8', textvariable=start_date)
start_date_entry.place(x=130, y=375)

start_date_btn = tk.Button(border=0, image=calendar_img, command=set_start_date)
start_date_btn.place(x=465, y=380)

tk.Label(window, text="End Date", font=f'{lato_font} 14', fg='#242424', bg='#ffffff').place(x=500, y=375)

end_date = tk.StringVar()
end_date_entry = tk.Entry(width=30, borderwidth=0, font=f'{lato_font} 14', bg='#f8f8f8', textvariable=end_date)
end_date_entry.place(x=600, y=375)

end_date_btn = tk.Button(border=0, image=calendar_img, command=set_end_date)
end_date_btn.place(x=935, y=380)

tk.Label(window, text="Address", font=f'{lato_font} 14', fg='#242424', bg='#ffffff').place(x=30, y=415)

sent_received_address = tk.StringVar()
sent_received_address_entry = tk.Entry(width=90, borderwidth=0, font=f'{lato_font} 14', bg='#f8f8f8', textvariable=sent_received_address) 
sent_received_address_entry.place(x=120, y=415)

include_ugl = tk.IntVar()
include_ugl_cb = tk.Checkbutton(window, bg='#ffffff', font=f'{lato_font} 14', text='Include unrealized gains or losses', variable=include_ugl)
include_ugl_cb.place(x=30, y=455)

send_received_status = tk.StringVar()
send_received_status_txt = tk.Entry(width=60, borderwidth=0, font=f'{lato_font} 12 italic', bg='#ffffff', textvariable=send_received_status) 
send_received_status_txt.place(x=750, y=455)

generate_sent_received_btn = tk.Button(text="   Generate Excel   ", borderwidth=0, font=f'{lato_font} 14', bg='#6030b1', fg='#ffffff', command=lambda:run_thread(generate_sent_received_history))
generate_sent_received_btn.place(x=1120, y=410)

tk.Label(window, text="Historical Price Data", font=f'{lato_font} 18', fg='#6030b1', bg='#ffffff').place(x=20, y=505)

tk.Label(window, text="Start Date", font=f'{lato_font} 14', fg='#242424', bg='#ffffff').place(x=30, y=555)

history_start_date = tk.StringVar()
history_start_date_entry = tk.Entry(width=30, borderwidth=0, font=f'{lato_font} 14', bg='#f8f8f8', textvariable=history_start_date)
history_start_date_entry.place(x=130, y=555)

history_start_date_entry_btn = tk.Button(border=0, image=calendar_img, command=set_history_start_date)
history_start_date_entry_btn.place(x=465, y=560)

tk.Label(window, text="End Date", font=f'{lato_font} 14', fg='#242424', bg='#ffffff').place(x=500, y=555)

history_end_date = tk.StringVar()
history_end_date_entry = tk.Entry(width=30, borderwidth=0, font=f'{lato_font} 14', bg='#f8f8f8', textvariable=history_end_date)
history_end_date_entry.place(x=600, y=555)

history_end_date_btn = tk.Button(border=0, image=calendar_img, command=set_history_end_date)
history_end_date_btn.place(x=935, y=560)

generate_history_btn = tk.Button(text="   Generate Excel   ", borderwidth=0, font=f'{lato_font} 14', bg='#6030b1', fg='#ffffff', command=lambda:run_thread(generate_historical_price_data))
generate_history_btn.place(x=1120, y=550)


MARKETS = ['kraken','gemini','bitfinex','binance-us','coinbase-pro']

current_market = tk.StringVar()
current_market.set(MARKETS[0])

tk.Label(window, text="Select Exchange", font=f'{lato_font} 14', fg='#242424', bg='#ffffff').place(x=30, y=600)

select_market = tk.OptionMenu(window, current_market, *MARKETS)
select_market.config(font=f'{lato_font} 14', border=0, bg='#f8f8f8')
select_market_menu = window.nametowidget(select_market.menuname)
select_market_menu.config(font=f'{lato_font} 14')
select_market.place(x=190, y=599)

history_status = tk.StringVar()
history_status_txt = tk.Entry(width=60, borderwidth=0, font=f'{lato_font} 12 italic', bg='#ffffff', textvariable=history_status) 
history_status_txt.place(x=750, y=600)

logo_img = tk.PhotoImage(file="res/logo_resized.png")
tk.Label(window, image=logo_img, bg='#ffffff').place(x=1075, y=650)

window.mainloop()





